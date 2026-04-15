#!/usr/bin/env python3
"""
Prepare JSON data files for the presentation from source databases.

Combines:
  1. regulators_performance_database (32 countries, 3M+ obs, recent data)
  2. old IBNET export (152 countries, 33K records, 1994-2022) as backup

Generates:
  - data/nrw_trends.json
  - data/metering_trends.json
  - data/cost_coverage_trends.json
  - data/coverage_map.json

Tariff timeseries already copied from ibnet-tariff-portal (timeseries.json).
"""

import sys
from pathlib import Path

import pandas as pd
import numpy as np

sys.path.insert(0, str(Path(__file__).parent))
from utils import (
    DATA_DIR, REG_DB, IBNET_GWI,
    load_country_region_mapping, write_json
)

# Country name -> ISO3 mapping for old IBNET (which uses full names)
COUNTRY_NAME_TO_ISO3 = {
    "Afghanistan": "AFG", "Albania": "ALB", "Algeria": "DZA",
    "American Samoa": "ASM", "Argentina": "ARG", "Armenia": "ARM",
    "Australia": "AUS", "Azerbaijan": "AZE", "Bahrain": "BHR",
    "Bangladesh": "BGD", "Belarus": "BLR", "Belgium": "BEL",
    "Benin": "BEN", "Bhutan": "BTN", "Bolivia": "BOL",
    "Bosnia and Herzegovina": "BIH", "Botswana": "BWA", "Brazil": "BRA",
    "Bulgaria": "BGR", "Burkina Faso": "BFA", "Burundi": "BDI",
    "Cambodia": "KHM", "Cameroon": "CMR", "Cape Verde": "CPV",
    "Central African Republic": "CAF", "Chad": "TCD", "Chile": "CHL",
    "China": "CHN", "Colombia": "COL", "Congo": "COG",
    "Congo, Dem. Rep.": "COD", "Cook Islands": "COK",
    "Costa Rica": "CRI", "Cote d'Ivoire": "CIV", "Croatia": "HRV",
    "Cyprus": "CYP", "Czech Republic": "CZE", "Denmark": "DNK",
    "Djibouti": "DJI", "Dominican Republic": "DOM", "Ecuador": "ECU",
    "Egypt": "EGY", "El Salvador": "SLV", "Eritrea": "ERI",
    "Ethiopia": "ETH", "Federated States Of Micronesia": "FSM",
    "Fiji": "FJI", "Finland": "FIN", "French Polynesia": "PYF",
    "Gabon": "GAB", "Georgia": "GEO", "Ghana": "GHA",
    "Guam": "GUM", "Guinea": "GIN", "Guinea-Bussau": "GNB",
    "Honduras": "HND", "Hungary": "HUN", "India": "IND",
    "Indonesia": "IDN", "Iraq": "IRQ", "Jordan": "JOR",
    "Kazakhstan": "KAZ", "Kenya": "KEN", "Kosovo": "XKX",
    "Kuwait": "KWT", "Kyrgyz Republic": "KGZ", "Lao PDR": "LAO",
    "Lesotho": "LSO", "Liberia": "LBR", "Libya": "LBY",
    "Lithuania": "LTU", "Macau, China": "MAC", "Madagascar": "MDG",
    "Malawi": "MWI", "Malaysia": "MYS", "Maldives": "MDV",
    "Mali": "MLI", "Malta": "MLT", "Marshall Islands": "MHL",
    "Mauritania": "MRT", "Mauritius": "MUS", "Mexico": "MEX",
    "Moldova": "MDA", "Mongolia": "MNG", "Montenegro": "MNE",
    "Mozambique": "MOZ", "Myanmar": "MMR", "Namibia": "NAM",
    "Nepal": "NPL", "Netherlands Antilles": "ANT",
    "New Caledonia": "NCL", "New Zealand": "NZL", "Nicaragua": "NIC",
    "Niger": "NER", "Nigeria": "NGA", "Niue": "NIU",
    "North Macedonia": "MKD", "Northern Mariana Islands": "MNP",
    "Norway": "NOR", "Oman": "OMN", "Pakistan": "PAK",
    "Palau": "PLW", "Panama": "PAN", "Papua New Guinea": "PNG",
    "Paraguay": "PRY", "Peru": "PER", "Philippines": "PHL",
    "Poland": "POL", "Portugal": "PRT", "Republic Of Kiribati": "KIR",
    "Republic Of Nauru": "NRU", "Romania": "ROU", "Russia": "RUS",
    "Rwanda": "RWA", "Samoa": "WSM", "Senegal": "SEN",
    "Serbia": "SRB", "Seychelles": "SYC", "Sierra Leone": "SLE",
    "Singapore": "SGP", "Slovakia": "SVK", "Solomon Islands": "SLB",
    "South Africa": "ZAF", "South Korea": "KOR", "Sri Lanka": "LKA",
    "Sudan": "SDN", "Sweden": "SWE", "Switzerland": "CHE",
    "Tajikistan": "TJK", "Tanzania": "TZA", "The Gambia": "GMB",
    "Togo": "TGO", "Tonga": "TON", "Tunisia": "TUN",
    "Turkey": "TUR", "Tuvalu": "TUV", "UK, England and Wales": "GBR",
    "UK, Scotland": "GBR", "Uganda": "UGA", "Ukraine": "UKR",
    "United States": "USA", "Uruguay": "URY", "Uzbekistan": "UZB",
    "Vanuatu": "VUT", "Venezuela": "VEN", "Vietnam": "VNM",
    "Wallis and Futuna": "WLF", "West Bank and Gaza": "PSE",
    "Yemen": "YEM", "Zambia": "ZMB", "Zimbabwe": "ZWE",
    "eSwatini": "SWZ",
}


def load_regulators_db():
    """Load the combined observations from regulators_performance_database."""
    path = REG_DB / "combined" / "observations.csv"
    print("Loading regulators DB...")
    df = pd.read_csv(path, dtype={"country_iso3": str, "indicator_id": str, "year": float, "value": float})
    print(f"  {len(df):,} observations, {df['country_iso3'].nunique()} countries")
    return df


def load_ibnet_backup():
    """Load old IBNET export and compute derived indicators."""
    path = IBNET_GWI / "old_IBNET_Export_RawData_20230310(Export) (1).csv"
    print("Loading IBNET backup...")
    cols = ["Country", "Year", "R_55_VOLUME_WATER_PRODUCED", "R_59_VOLUME_WATER_SOLD",
            "R_53_POP_CONNECTIONS_OPERATING_METER", "R_41_POP_CONNECTIONS_YEAR_END",
            "R_90_TOTAL_OPERATING_REVENUE", "R_94_TOTAL_OPERATING_EXPENSES"]
    df = pd.read_csv(path, encoding="latin-1", usecols=cols)
    # Convert numeric columns — strip commas first (old IBNET has "57,237,112.48" format)
    for c in cols:
        if c not in ("Country", "Year"):
            df[c] = df[c].astype(str).str.replace(",", "", regex=False)
            df[c] = pd.to_numeric(df[c], errors="coerce")
    df["country_iso3"] = df["Country"].map(COUNTRY_NAME_TO_ISO3)
    df["year"] = pd.to_numeric(df["Year"], errors="coerce")
    print(f"  {len(df):,} records, {df['country_iso3'].nunique()} countries mapped")

    rows = []

    # NRW = (Produced - Sold) / Produced * 100
    nrw_mask = df["R_55_VOLUME_WATER_PRODUCED"].notna() & df["R_59_VOLUME_WATER_SOLD"].notna()
    nrw_df = df[nrw_mask].copy()
    nrw_df["value"] = (nrw_df["R_55_VOLUME_WATER_PRODUCED"] - nrw_df["R_59_VOLUME_WATER_SOLD"]) / nrw_df["R_55_VOLUME_WATER_PRODUCED"] * 100
    nrw_df["indicator_id"] = "OPS_NRW_PCT"
    rows.append(nrw_df[["country_iso3", "year", "value", "indicator_id"]])

    # Metering = Metered connections / Total connections * 100
    meter_mask = df["R_53_POP_CONNECTIONS_OPERATING_METER"].notna() & df["R_41_POP_CONNECTIONS_YEAR_END"].notna()
    meter_df = df[meter_mask].copy()
    meter_df["value"] = meter_df["R_53_POP_CONNECTIONS_OPERATING_METER"] / meter_df["R_41_POP_CONNECTIONS_YEAR_END"] * 100
    meter_df["indicator_id"] = "OPS_METERING_PCT"
    rows.append(meter_df[["country_iso3", "year", "value", "indicator_id"]])

    # Cost coverage = Revenue / Expenses * 100
    fin_mask = df["R_90_TOTAL_OPERATING_REVENUE"].notna() & df["R_94_TOTAL_OPERATING_EXPENSES"].notna()
    fin_df = df[fin_mask & (df["R_94_TOTAL_OPERATING_EXPENSES"] > 0)].copy()
    fin_df["value"] = fin_df["R_90_TOTAL_OPERATING_REVENUE"] / fin_df["R_94_TOTAL_OPERATING_EXPENSES"] * 100
    fin_df["indicator_id"] = "FIN_COST_COVERAGE"
    rows.append(fin_df[["country_iso3", "year", "value", "indicator_id"]])

    result = pd.concat(rows, ignore_index=True)
    result["source"] = "ibnet"
    print(f"  Derived {len(result):,} indicator observations")
    return result


def load_new_ibnet():
    """Load NewIBNET.xlsx (2021-2023 data) and compute NRW."""
    path = IBNET_GWI / "NewIBNET.xlsx"
    print("Loading NewIBNET...")
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Sheet1"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    rows = []
    for r in data_rows:
        iso3 = r[0]
        try:
            year = int(float(r[4])) if r[4] else None
            prod = float(r[10]) if r[10] else None
            billed = float(r[12]) if r[12] else None
        except (ValueError, TypeError):
            continue
        if not (iso3 and year and prod and prod > 0 and billed is not None):
            continue
        nrw = (prod - billed) / prod * 100
        if 5 <= nrw <= 100:
            rows.append({"country_iso3": iso3, "year": year, "value": nrw, "indicator_id": "OPS_NRW_PCT"})

    result = pd.DataFrame(rows)
    result["source"] = "new_ibnet"
    print(f"  {len(result):,} NRW observations, {result['country_iso3'].nunique()} countries")
    return result


    # Normalize region names: WB API uses "&" but chart JS uses "and"
REGION_NAME_NORMALIZE = {
    "East Asia & Pacific": "East Asia and Pacific",
    "Europe & Central Asia": "Europe and Central Asia",
    "Latin America & Caribbean": "Latin America and the Caribbean",
    "Middle East, North Africa, Afghanistan & Pakistan": "Middle East, North Africa, Afghanistan and Pakistan",
}


def compute_regional_trends(df, indicator_id, country_region_map, min_countries=2, min_year=2005, rolling_window=0):
    """
    Compute median-of-medians by region and year.
    min_year=2005: earlier years have sparse, inconsistent data across regions.
    min_countries=2: require at least 2 countries per region-year for a stable median.
    Global trend requires min 3 countries (hardcoded below).
    Country median first, then regional median (prevents large-country domination).
    rolling_window: if >0, apply N-year rolling median smoothing to reduce composition effects.
    """
    ind = df[df["indicator_id"] == indicator_id].copy()
    if len(ind) == 0:
        print(f"  WARNING: No data for {indicator_id}")
        return {}

    ind["region"] = ind["country_iso3"].map(country_region_map)
    ind = ind.dropna(subset=["region", "value", "year"])
    ind = ind[ind["year"] >= min_year]

    # Remove outliers
    # NRW: values in range (0, 1) are decimal-format (e.g. 0.41 means 41%) — rescue them
    if "NRW" in indicator_id:
        decimal_mask = (ind["value"] > 0) & (ind["value"] < 1.0)
        n_rescued = decimal_mask.sum()
        if n_rescued > 0:
            ind.loc[decimal_mask, "value"] = ind.loc[decimal_mask, "value"] * 100
            print(f"  NRW: rescued {n_rescued} decimal-format values (multiplied by 100)")
        ind = ind[(ind["value"] >= 5) & (ind["value"] <= 100)]
    elif "METERING" in indicator_id:
        ind = ind[(ind["value"] >= 0) & (ind["value"] <= 100)]
    elif "COST_COVERAGE" in indicator_id:
        ind = ind[(ind["value"] > 0) & (ind["value"] <= 500)]

    # Step 1: country medians per year
    country_medians = ind.groupby(["region", "country_iso3", "year"])["value"].median().reset_index()

    # Step 2: regional medians
    result = {}
    for region in sorted(country_medians["region"].unique()):
        reg_data = country_medians[country_medians["region"] == region]
        yearly = reg_data.groupby("year").agg(
            median=("value", "median"),
            n_countries=("country_iso3", "nunique"),
            n_obs=("value", "count")
        ).reset_index()
        yearly = yearly[yearly["n_countries"] >= min_countries].sort_values("year")
        if rolling_window > 0 and len(yearly) >= rolling_window:
            yearly["median"] = yearly["median"].rolling(rolling_window, center=True, min_periods=1).median()
        region_key = REGION_NAME_NORMALIZE.get(region, region)
        if len(yearly) > 0:
            result[region_key] = [
                {"year": int(row["year"]), "median": round(float(row["median"]), 2),
                 "count": int(row["n_obs"]), "nCountries": int(row["n_countries"])}
                for _, row in yearly.iterrows()
            ]

    # Global
    global_yearly = country_medians.groupby("year").agg(
        median=("value", "median"),
        n_countries=("country_iso3", "nunique")
    ).reset_index()
    global_yearly = global_yearly[global_yearly["n_countries"] >= 3].sort_values("year")
    if rolling_window > 0 and len(global_yearly) >= rolling_window:
        global_yearly["median"] = global_yearly["median"].rolling(rolling_window, center=True, min_periods=1).median()
    if len(global_yearly) > 0:
        result["Global"] = [
            {"year": int(row["year"]), "median": round(float(row["median"]), 2),
             "count": int(row["n_countries"])}
            for _, row in global_yearly.iterrows()
        ]

    return result


def build_coverage_map(country_region_map):
    """Build country-level coverage data for the map visualization."""
    summary_path = REG_DB / "combined" / "country_summary.csv"
    summary = pd.read_csv(summary_path)

    coverage = []
    for _, row in summary.iterrows():
        iso3 = row["country_iso3"]
        coverage.append({
            "iso3": iso3,
            "name": row["country_name"],
            "region": country_region_map.get(iso3, "Unknown"),
            "nUtilities": int(row["n_utilities"]),
            "nIndicators": int(row["n_indicators"]),
            "nObservations": int(row["n_observations"]),
            "yearMin": int(row["year_min"]),
            "yearMax": int(row["year_max"]),
            "sources": row["sources"],
            "tier": "regulators_db"
        })

    stats = {
        "totalCountries": len(coverage),
        "totalUtilities": sum(c["nUtilities"] for c in coverage),
        "totalObservations": sum(c["nObservations"] for c in coverage),
    }
    return {"countries": coverage, "stats": stats}


def clean_tariff_data():
    """
    Clean the portal tariff timeseries.json:
    - Remove years > 2024 (test/future data)
    - Remove region-year combos with < 5 observations
    - Remove individual median values < $0.10/15m³ (decimal errors)
    """
    import json
    portal_path = Path("/Users/anademenezes/Documents/ibnet-tariff-portal/web/public/data/timeseries.json")
    print("\nCleaning tariff data from portal...")
    with open(portal_path) as f:
        raw = json.load(f)

    # Map IBNET region codes to full WB region names (matching hero-chart.js)
    IBNET_CODE_TO_FULL = {
        "AFR2": "Sub-Saharan Africa",
        "EAP": "East Asia & Pacific",
        "ECA": "Europe & Central Asia",
        "LAC": "Latin America & Caribbean",
        "MNA": "Middle East, North Africa, Afghanistan & Pakistan",
        "SAR": "South Asia",
        "Global": "Global",
    }

    cleaned = {}
    for region_code, entries in raw.items():
        filtered = []
        for e in entries:
            year = e.get("year", 9999)
            median15 = e.get("median15m3", 0)
            count = e.get("count", 0)
            # Remove future years, tiny observation counts, and obvious decimal errors
            if year > 2024:
                continue
            if count < 5:
                continue
            if median15 is not None and median15 < 0.50:
                print(f"    Dropped {region_code} {year}: ${median15:.2f}/15m³ (likely currency error)")
                continue
            filtered.append(e)
        # Use full region name as key (matching hero-chart.js expectations)
        full_name = IBNET_CODE_TO_FULL.get(region_code, region_code)
        if filtered:
            cleaned[full_name] = filtered
            years = [e["year"] for e in filtered]
            print(f"  {full_name}: {min(years)}-{max(years)}, {len(filtered)} years (was {len(entries)})")

    write_json(cleaned, DATA_DIR / "timeseries.json")
    return cleaned


def main():
    print("=" * 60)
    print("Preparing presentation data")
    print("=" * 60)

    # Load region mapping
    print("\nLoading country-region mapping...")
    country_region_map = load_country_region_mapping()
    print(f"  {len(country_region_map)} countries mapped to regions")

    # Load all three data sources
    print("\nLoading data sources...")
    reg_df = load_regulators_db()
    ibnet_df = load_ibnet_backup()
    new_ibnet_df = load_new_ibnet()

    # Combine: regulators DB > NewIBNET > old IBNET (priority order for dedup)
    # Tag regulators data
    reg_subset = reg_df[reg_df["indicator_id"].isin(["OPS_NRW_PCT", "OPS_METERING_PCT", "FIN_COST_COVERAGE"])].copy()
    reg_subset["source"] = "regulators"

    # Build priority keys: regulators first
    reg_keys = set(zip(reg_subset["country_iso3"], reg_subset["year"].astype(int), reg_subset["indicator_id"]))

    # Dedup NewIBNET against regulators
    new_ibnet_before = len(new_ibnet_df)
    new_ibnet_filtered = new_ibnet_df[~new_ibnet_df.apply(
        lambda r: (r["country_iso3"], int(r["year"]) if pd.notna(r["year"]) else 0, r["indicator_id"]) in reg_keys,
        axis=1
    )]
    print(f"  NewIBNET dedup: {new_ibnet_before - len(new_ibnet_filtered):,} dropped (overlap with regulators), {len(new_ibnet_filtered):,} kept")

    # Build combined priority keys (regulators + NewIBNET)
    combined_keys = reg_keys.copy()
    for _, r in new_ibnet_filtered.iterrows():
        if pd.notna(r["year"]):
            combined_keys.add((r["country_iso3"], int(r["year"]), r["indicator_id"]))

    # Dedup old IBNET against regulators + NewIBNET
    ibnet_before = len(ibnet_df)
    ibnet_filtered = ibnet_df[~ibnet_df.apply(
        lambda r: (r["country_iso3"], int(r["year"]) if pd.notna(r["year"]) else 0, r["indicator_id"]) in combined_keys,
        axis=1
    )]
    ibnet_dropped = ibnet_before - len(ibnet_filtered)
    print(f"  Old IBNET dedup: {ibnet_dropped:,} dropped (overlap with regulators + NewIBNET), {len(ibnet_filtered):,} kept")

    cols = ["country_iso3", "year", "value", "indicator_id", "source"]
    combined = pd.concat([reg_subset[cols], new_ibnet_filtered[cols], ibnet_filtered[cols]],
                         ignore_index=True)

    print(f"\nCombined dataset: {len(combined):,} records")
    for ind in ["OPS_NRW_PCT", "OPS_METERING_PCT", "FIN_COST_COVERAGE"]:
        sub = combined[combined["indicator_id"] == ind]
        n_reg = len(sub[sub["source"] == "regulators"])
        n_new = len(sub[sub["source"] == "new_ibnet"])
        n_ibnet = len(sub[sub["source"] == "ibnet"])
        n_countries = sub["country_iso3"].nunique()
        print(f"  {ind}: {n_reg:,} regulators + {n_new:,} NewIBNET + {n_ibnet:,} old IBNET = {len(sub):,} total ({n_countries} countries)")

    # Generate trends
    for indicator, filename, label, rw in [
        ("OPS_NRW_PCT", "nrw_trends.json", "NRW Trends", 3),
        ("OPS_METERING_PCT", "metering_trends.json", "Metering Trends", 0),
        ("FIN_COST_COVERAGE", "cost_coverage_trends.json", "Cost Coverage Trends", 0),
    ]:
        print(f"\n--- {label} ---")
        trends = compute_regional_trends(combined, indicator, country_region_map, rolling_window=rw)
        for region, data in sorted(trends.items()):
            years = [d["year"] for d in data]
            print(f"  {region}: {min(years)}-{max(years)}, {len(data)} years")
        write_json(trends, DATA_DIR / filename)

    # Coverage map
    print("\n--- Coverage Map ---")
    coverage = build_coverage_map(country_region_map)
    print(f"  {coverage['stats']['totalCountries']} countries, "
          f"{coverage['stats']['totalUtilities']:,} utilities, "
          f"{coverage['stats']['totalObservations']:,} observations")
    write_json(coverage, DATA_DIR / "coverage_map.json")

    # Clean tariff data
    clean_tariff_data()

    print("\n" + "=" * 60)
    print("Done! All data files written to data/")
    print("=" * 60)


if __name__ == "__main__":
    main()
